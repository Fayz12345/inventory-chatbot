# Load `.env` BEFORE any other module imports — ecommerce.config reads from
# os.environ at import time, so dotenv must populate it first.
from dotenv import load_dotenv
load_dotenv()

from flask import Flask, render_template, request, session, redirect, url_for, jsonify, Response
from werkzeug.security import check_password_hash
from functools import wraps
import csv
import io
import json
import os
import secrets as _secrets
import time
from urllib import error as urlerror, parse as urlparse, request as urlrequest
import pyodbc
import anthropic
import config
import users_db
import admin_audit
import chat_sql
import chat_log
import roles

CHAT_SQL_MODEL = getattr(config, "CHAT_SQL_MODEL", "claude-sonnet-4-6")
CHAT_ANSWER_MODEL = getattr(config, "CHAT_ANSWER_MODEL", "claude-haiku-4-5-20251001")
CHAT_ROW_CAP = 50
EXPORT_ROW_CAP = 5000
CHAT_HISTORY_TURNS = 6

_anthropic = None
def _anthropic_client():
    """One shared client with a timeout so a hung API call can't pin a worker."""
    global _anthropic
    if _anthropic is None:
        _anthropic = anthropic.Anthropic(
            api_key=config.ANTHROPIC_API_KEY, timeout=30, max_retries=2
        )
    return _anthropic


chatbot_app = Flask(__name__)
chatbot_app.secret_key = config.SECRET_KEY


def require_module(module):
    """Decorator: redirect to /home if the session role lacks the module."""
    def deco(fn):
        @wraps(fn)
        def wrapper(*a, **kw):
            if not session.get('logged_in'):
                return redirect(url_for('login'))
            role = roles.effective_role(session.get('role'), session.get('is_admin'))
            if not roles.role_allows(role, module):
                return redirect(url_for('home'))
            return fn(*a, **kw)
        return wrapper
    return deco


def _perms():
    """Return a perms dict for the current session role (for nav gating)."""
    role = roles.effective_role(session.get('role'), session.get('is_admin'))
    return {m: roles.role_allows(role, m) for m in roles.MODULES}

# Initialise local SQLite user database
users_db.init_db()
users_db.seed_admin_if_empty()
admin_audit.init_db()
chat_log.init_db()

# Register blueprints
from ecommerce.approval import approval_bp
from analytics.routes import analytics_bp
from billing.routes import billing_bp
chatbot_app.register_blueprint(approval_bp)
chatbot_app.register_blueprint(analytics_bp)
chatbot_app.register_blueprint(billing_bp)

# eBay Marketplace Account Deletion (MAD) webhook helpers (public route below).
from ecommerce.notifications import ebay_deletion


# --- CSRF protection ---
@chatbot_app.before_request
def _csrf_guard():
    if session.get('logged_in') and not session.get('csrf_token'):
        session['csrf_token'] = _secrets.token_urlsafe(32)
    protected = request.method == 'POST' and (
        request.path.startswith('/admin/') or request.path.startswith('/ask') or request.path.startswith('/profile'))
    if protected:
        token = session.get('csrf_token')
        # reject when there is no session token (avoids a None==None pass for an
        # unauthenticated POST to a protected path) or the header doesn't match
        if not token or request.headers.get('X-CSRF-Token') != token:
            return jsonify({'ok': False, 'error': 'Invalid CSRF token'}), 403


@chatbot_app.context_processor
def _inject_csrf():
    return {'csrf_token': session.get('csrf_token', '')}


@chatbot_app.context_processor
def _inject_perms():
    """Make `perms` available to EVERY template so the nav gates by role
    consistently — previously only home/chat/ecommerce passed it, so admin and
    other pages fell through `perms is not defined` and showed all tabs."""
    return {'perms': _perms()}


# --- eBay Marketplace Account Deletion (MAD) webhook ---
# PUBLIC endpoint (no login) required to enable the production eBay keyset. It is
# CSRF-exempt because _csrf_guard only protects /admin/, /ask, /profile POSTs, and
# it lives at the top level (NOT under the ecommerce blueprint, which force-logins
# every route). Exact path with no trailing slash so eBay's ?challenge_code= GET
# isn't 30x-redirected (which eBay rejects). See ecommerce/notifications/ebay_deletion.py.
@chatbot_app.route('/ebay/account-deletion', methods=['GET', 'POST'])
def ebay_account_deletion():
    if request.method == 'GET':
        code = request.args.get('challenge_code')
        if not code:
            return jsonify({'error': 'missing challenge_code'}), 400
        # 200 + application/json {"challengeResponse": sha256(code+token+endpoint)}
        return jsonify({'challengeResponse': ebay_deletion.challenge_response(code)}), 200
    # POST: always ack 200; email + log on a valid payload; never 5xx back to eBay.
    try:
        ebay_deletion.handle_notification(request.get_json(silent=True) or {})
    except Exception:
        chatbot_app.logger.exception("eBay deletion notification handling failed")
    return '', 200


# --- Public privacy policy ---
# PUBLIC page (no login) linked from eBay's OAuth consent screen at
# https://ai.bridge-renew.net/privacy — it must return 200 with a real policy.
# Top-level route (NOT under the ecommerce blueprint, which force-logins every
# route) and no login decorator; it's a standalone legal page with no app nav.
@chatbot_app.route('/privacy', methods=['GET'])
def privacy():
    return render_template('privacy.html')


# --- Database connection ---
def get_db_connection():
    conn = pyodbc.connect(
        f"DRIVER={{ODBC Driver 18 for SQL Server}};"
        f"SERVER={config.DB_SERVER};"
        f"DATABASE={config.DB_NAME};"
        f"UID={config.DB_USER};"
        f"PWD={config.DB_PASSWORD};"
        f"TrustServerCertificate=yes;"
    )
    return conn

# --- Table schema sent to Claude so it knows what to query ---
TABLE_SCHEMA = """
You have access to these tables. Choose the ONE that matches the question.

TABLE 1 — ReportingInventoryFlat (current IN-STOCK inventory; one row per device):
- ESN (nvarchar): unique device identifier / IMEI
- ProjectName (nvarchar): the project the device belongs to
- ProjectTag (nvarchar): project tag or label
- ReceiveDate (datetime): date the device was received
- Product_Place (nvarchar): physical location of the device e.g. 'Product Room'
- Manufacturer (nvarchar): device manufacturer e.g. 'Apple', 'Samsung'
- Model (nvarchar): device model e.g. 'iPhone 14 Pro'
- Colour (nvarchar): device colour
- Grade (nvarchar): device grade
- Received_Grade (nvarchar): grade at time of receiving
- DeviceCost (decimal): cost of the device
- Function_Test_Created (nvarchar): date function test was completed
- Grading_Created (nvarchar): date grading was completed
- LastRefreshed (datetime): when the table was last refreshed
Contains in-stock devices only.

TABLE 2 — EcommerceListingsLog (marketplace listings WE HAVE POSTED; one row per listing):
- Manufacturer, Model, Colour, Grade, Quantity
- Platform (nvarchar): the marketplace — 'eBay CA', 'Best Buy CA', 'Amazon CA', 'Reebelo CA'
- ListingPrice (decimal): the price we listed it at
- FloorPriceAtListing (decimal): the competitor floor price at listing time
- PlatformListingID (nvarchar): the marketplace's listing id ('manual' if listed manually)
- Status (nvarchar): 'active', 'ended', or 'sold'
- CreatedAt (datetime): when we posted the listing
- EndedAt (datetime): when the listing ended/sold (NULL while active)
- ApprovedBy (nvarchar): the user who approved it
Use for "what / how many have we listed / posted to a marketplace".

TABLE 3 — EcommercePricingRecommendation (weekly pricing recommendations; one row per device variant per run):
- BatchID (int): the pricing-run batch (join EcommercePricingBatch.ID for its date)
- Manufacturer, Model, Colour, Grade, Quantity
- RecommendedMarketplace (nvarchar): marketplace with the best price — 'Amazon CA','eBay CA','Best Buy CA','Reebelo CA'
- RecommendedPrice (decimal): the recommended list price
- AmazonFloor, EbayFloor, BestBuyFloor, ReebeloFloor (decimal): per-marketplace competitor floors
- DeviceCost (decimal): our cost
- MarginOK (bit): 1 if the margin is acceptable, 0 if skipped
- SkipReason (nvarchar): why it was skipped (when MarginOK = 0)
- Decision (nvarchar): 'approved', 'rejected', or NULL (NULL = pending / not yet decided)
- DecidedAt (datetime): when it was approved/rejected
Use for recommendations / approved / rejected / pending / recommended prices / margins.

TABLE 4 — EcommercePricingBatch (one row per weekly pricing run):
- ID (int): batch id (matches EcommercePricingRecommendation.BatchID)
- CreatedAt (datetime): when the batch ran
- Status (nvarchar): batch status
"""

SYSTEM_PROMPT = f"""
You are a data assistant for a phone-refurbishment business. You answer questions by
generating a single T-SQL SELECT query for Microsoft SQL Server 2019 against the tables below.

{TABLE_SCHEMA}

Rules:
- Generate T-SQL for SQL Server. Use TOP (e.g. SELECT TOP 100 ...) to limit rows.
  NEVER use LIMIT — it is not valid T-SQL and will fail.
- For any question that lists rows (not an aggregate), add TOP 100.
- Only generate SELECT queries. Never INSERT/UPDATE/DELETE/DROP or any write.
- Only query the tables described above. Pick the ONE table that matches the question:
  in-stock inventory -> ReportingInventoryFlat; listings we've posted to a marketplace ->
  EcommerceListingsLog; pricing recommendations (approved/rejected/pending, recommended
  prices, margins) -> EcommercePricingRecommendation (join EcommercePricingBatch on BatchID
  only when you need the batch date). Do NOT join ReportingInventoryFlat with the ecommerce
  tables — they share no key.
- Overall device SALES ("how many sold last month" across all channels) are NOT in these
  tables -> respond UNABLE_TO_ANSWER. (EcommerceListingsLog.Status = 'sold' only reflects
  marketplace listings we explicitly marked sold.)
- Function_Test_Created, Grading_Created, Grade and Received_Grade are nvarchar (text)
  columns. Dates in them are 'YYYY-MM-DD...' strings — compare as strings or wrap with
  TRY_CONVERT(date, col); do not assume they are real datetimes.
- Match manufacturer/model loosely with LIKE and wildcards (e.g. Model LIKE '%iPhone 14%').
- Return ONLY the raw SQL query: no explanation, no markdown, no code fences.
- If the question cannot be answered from these tables, respond with: UNABLE_TO_ANSWER

Examples:
Q: how many Samsung devices are in stock?
SQL: SELECT COUNT(*) FROM ReportingInventoryFlat WHERE Manufacturer LIKE '%Samsung%'

Q: list 10 iPhones in the Product Room
SQL: SELECT TOP 10 ESN, Model, Colour, Grade FROM ReportingInventoryFlat WHERE Model LIKE '%iPhone%' AND Product_Place = 'Product Room'

Q: how many devices of each grade?
SQL: SELECT Grade, COUNT(*) AS cnt FROM ReportingInventoryFlat GROUP BY Grade ORDER BY cnt DESC

Q: how many devices have we listed on each marketplace?
SQL: SELECT Platform, COUNT(*) AS listings FROM EcommerceListingsLog GROUP BY Platform ORDER BY listings DESC

Q: how many Samsung devices have we posted to the ecommerce marketplaces?
SQL: SELECT COUNT(*) FROM EcommerceListingsLog WHERE Manufacturer LIKE '%Samsung%'

Q: how many pricing recommendations are pending?
SQL: SELECT COUNT(*) FROM EcommercePricingRecommendation WHERE Decision IS NULL

Q: what is the recommended price for the iPhone 14 Pro?
SQL: SELECT TOP 100 Model, Colour, Grade, RecommendedMarketplace, RecommendedPrice FROM EcommercePricingRecommendation WHERE Model LIKE '%iPhone 14 Pro%'

Q: what is the weather today?
SQL: UNABLE_TO_ANSWER
"""

# --- Microsoft Graph Email ---
_graph_token_cache = {"access_token": "", "expires_at": 0.0}

def _graph_access_token():
    now = time.time()
    if _graph_token_cache["access_token"] and _graph_token_cache["expires_at"] - 60 > now:
        return _graph_token_cache["access_token"]
    token_url = f"https://login.microsoftonline.com/{config.M365_TENANT_ID}/oauth2/v2.0/token"
    payload = urlparse.urlencode({
        "client_id": config.M365_CLIENT_ID,
        "client_secret": config.M365_CLIENT_SECRET,
        "scope": "https://graph.microsoft.com/.default",
        "grant_type": "client_credentials",
    }).encode("utf-8")
    req = urlrequest.Request(token_url, data=payload,
                             headers={"Content-Type": "application/x-www-form-urlencoded"}, method="POST")
    with urlrequest.urlopen(req, timeout=30) as resp:
        token_data = json.load(resp)
    _graph_token_cache["access_token"] = token_data["access_token"]
    _graph_token_cache["expires_at"] = now + max(int(token_data.get("expires_in", 0)), 300)
    return token_data["access_token"]

def send_invite_email(email, username, token):
    link = f"{config.APP_URL}/set-password/{token}"
    html = f"""\
<div style="font-family:Arial,sans-serif;max-width:520px;margin:0 auto;padding:30px">
  <div style="background:#2563eb;color:#fff;padding:18px 24px;border-radius:8px 8px 0 0">
    <h2 style="margin:0;font-size:20px">Bridge Platform</h2>
  </div>
  <div style="background:#fff;border:1px solid #e5e7eb;border-top:none;padding:28px 24px;border-radius:0 0 8px 8px">
    <p style="margin:0 0 16px;color:#333;font-size:15px">Hi <strong>{username}</strong>,</p>
    <p style="margin:0 0 16px;color:#333;font-size:15px">
      You've been invited to <strong>Bridge Platform</strong>. Click the button below to set your password and get started.
    </p>
    <p style="margin:0 0 16px;color:#333;font-size:15px">
      After setting your password you can sign in with this email address <em>or</em> your username <strong>{username}</strong>.
    </p>
    <div style="text-align:center;margin:28px 0">
      <a href="{link}"
         style="display:inline-block;padding:12px 32px;background:#2563eb;color:#fff;text-decoration:none;
                border-radius:6px;font-weight:bold;font-size:15px">
        Set Your Password
      </a>
    </div>
    <p style="margin:0 0 8px;color:#666;font-size:13px">Or copy this link into your browser:</p>
    <p style="margin:0 0 20px;color:#2563eb;font-size:13px;word-break:break-all">{link}</p>
    <p style="margin:0;color:#999;font-size:12px">This link expires in 7 days.</p>
  </div>
</div>"""
    message = {
        "subject": "You've been invited to Bridge Platform",
        "body": {"contentType": "HTML", "content": html},
        "toRecipients": [{"emailAddress": {"address": email}}],
    }
    graph_url = f"https://graph.microsoft.com/v1.0/users/{urlparse.quote(config.M365_SENDER)}/sendMail"
    req = urlrequest.Request(
        graph_url,
        data=json.dumps({"message": message, "saveToSentItems": True}).encode("utf-8"),
        headers={"Authorization": f"Bearer {_graph_access_token()}", "Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urlrequest.urlopen(req, timeout=30):
            return
    except urlerror.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"Graph sendMail failed: {body}") from exc


# --- Run SQL query safely ---
def run_query(sql):
    try:
        safe_sql = chat_sql.validate_sql(sql)
    except chat_sql.SqlValidationError as e:
        return None, str(e)
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(safe_sql)
        fetched = cursor.fetchmany(CHAT_ROW_CAP + 1)
        truncated = len(fetched) > CHAT_ROW_CAP
        fetched = fetched[:CHAT_ROW_CAP]
        columns = [col[0] for col in cursor.description]
        conn.close()
        return {'columns': columns,
                'rows': [list(r) for r in fetched],
                'truncated': truncated}, None
    except Exception as e:
        return None, str(e)


def run_query_raw(sql):
    try:
        safe_sql = chat_sql.validate_sql(sql)
        conn = get_db_connection(); cur = conn.cursor(); cur.execute(safe_sql)
        cols = [c[0] for c in cur.description]; rows = [list(r) for r in cur.fetchall()]
        conn.close()
        return {'columns': cols, 'rows': rows}, None
    except Exception as e:
        return None, str(e)

def _sanitize_history(raw):
    """Accept only well-formed {role in (user,assistant), content:str} items,
    keep the last CHAT_HISTORY_TURNS*2, and coerce to plain strings."""
    if not isinstance(raw, list):
        return []
    clean = []
    for m in raw[-(CHAT_HISTORY_TURNS * 2):]:
        if isinstance(m, dict) and m.get("role") in ("user", "assistant") and isinstance(m.get("content"), str):
            clean.append({"role": m["role"], "content": m["content"][:4000]})
    return clean

# --- Ask Claude to generate SQL ---
CHAT_MAX_RETRIES = 2

def generate_sql(messages):
    message = _anthropic_client().messages.create(
        model=CHAT_SQL_MODEL, max_tokens=500, system=SYSTEM_PROMPT, messages=messages)
    return message.content[0].text.strip(), message.usage

# --- Format result into a readable answer ---
ANSWER_SYSTEM = ("You are a helpful inventory assistant. Given a SQL query result, answer the "
                 "user's question in plain English. Be concise and direct. If it's a count or "
                 "sum, state the number clearly. When listing multiple items, use short bullet "
                 "points (lines starting with '- ').")

# Shown to the user in place of a raw DB/SQL error string (which stays in the log only).
FRIENDLY_ERROR = "I couldn't run a query for that — try rephrasing your question."


def _answer_prompt(sql, data, user_question, truncated, total_rows):
    note = ""
    if truncated:
        note = (f"\nNOTE: results were truncated to the first {CHAT_ROW_CAP} of "
                f"{total_rows if total_rows is not None else 'many'} rows. Do NOT "
                f"claim this is the complete set; say it's a sample.")
    return (f"Question: {user_question}\nSQL used: {sql}\nColumns: {data['columns']}\n"
            f"Data: {str(data['rows'])}{note}\n\nAnswer the question in plain English.")


def format_answer(sql, data, user_question, truncated=False, total_rows=None):
    message = _anthropic_client().messages.create(
        model=CHAT_ANSWER_MODEL, max_tokens=500, system=ANSWER_SYSTEM,
        messages=[{"role": "user", "content": _answer_prompt(sql, data, user_question, truncated, total_rows)}],
    )
    return message.content[0].text.strip()


def format_answer_stream(sql, data, user_question, truncated, total_rows, usage_out):
    """Yield the answer text chunk-by-chunk (streaming). Populates usage_out with
    the final token counts once the stream completes."""
    with _anthropic_client().messages.stream(
        model=CHAT_ANSWER_MODEL, max_tokens=500, system=ANSWER_SYSTEM,
        messages=[{"role": "user", "content": _answer_prompt(sql, data, user_question, truncated, total_rows)}],
    ) as stream:
        for text in stream.text_stream:
            yield text
        final = stream.get_final_message()
        usage_out['input_tokens'] = getattr(final.usage, 'input_tokens', 0)
        usage_out['output_tokens'] = getattr(final.usage, 'output_tokens', 0)


def _resolve_stream(user_question, history):
    """Generator over the SQL gen + self-correcting retry + run pipeline. Yields
    ``{"stage": name}`` progress events, then a final ``{"result": {...}, "in_tok",
    "out_tok", "retries"}``. `result["kind"]` is one of unable / error / empty / ok.
    Reused by both /ask (drain for the result) and /ask/stream (relay the stages)."""
    in_tok = out_tok = retries = 0
    messages = history + [{"role": "user", "content": user_question}]
    sql = data = error = None
    for attempt in range(CHAT_MAX_RETRIES + 1):
        yield {"stage": "writing_query" if attempt == 0 else "retrying"}
        sql, _usage = generate_sql(messages)
        in_tok += getattr(_usage, 'input_tokens', 0)
        out_tok += getattr(_usage, 'output_tokens', 0)
        if sql == 'UNABLE_TO_ANSWER':
            yield {"result": {"kind": "unable"}, "in_tok": in_tok, "out_tok": out_tok, "retries": retries}
            return
        yield {"stage": "running"}
        data, error = run_query(sql)
        if not error:
            break
        if attempt < CHAT_MAX_RETRIES:
            retries += 1
        messages.append({"role": "assistant", "content": sql})
        messages.append({"role": "user",
                         "content": f"That query failed with error: {error}. Return a corrected single T-SQL SELECT only."})
    if error:
        yield {"result": {"kind": "error", "sql": sql, "error": error},
               "in_tok": in_tok, "out_tok": out_tok, "retries": retries}
        return
    if not data['rows']:
        yield {"result": {"kind": "empty", "sql": sql, "columns": data['columns']},
               "in_tok": in_tok, "out_tok": out_tok, "retries": retries}
        return
    total_rows = len(data['rows'])
    if data.get('truncated'):
        count_data, _ = run_query_raw(chat_sql.build_count_query(sql))
        total_rows = count_data['rows'][0][0] if count_data else None
    yield {"result": {"kind": "ok", "sql": sql, "data": data,
                      "truncated": bool(data.get('truncated')), "total_rows": total_rows},
           "in_tok": in_tok, "out_tok": out_tok, "retries": retries}


def _resolve_and_run(user_question, history):
    """Non-streaming wrapper: drain _resolve_stream and return the final tuple."""
    final = None
    for item in _resolve_stream(user_question, history):
        if "result" in item:
            final = item
    return final["result"], final["in_tok"], final["out_tok"], final["retries"]


def _log_ask(username, question, t0, *, ok, sql=None, error=None, row_count=None,
             retries=0, in_tok=0, out_tok=0):
    try:
        chat_log.log_query(username=username, question=question, sql=sql, ok=ok, error=error,
                           row_count=row_count, retries=retries,
                           latency_ms=int((time.time() - t0) * 1000),
                           input_tokens=in_tok, output_tokens=out_tok)
    except Exception:
        pass

# --- Routes ---
@chatbot_app.route('/', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = users_db.authenticate(username, password)
        if user:
            session['logged_in'] = True
            session['username'] = user['username']
            session['role'] = user.get('role') or ('admin' if user['is_admin'] else 'user')
            session['is_admin'] = bool(user['is_admin'])
            return redirect(url_for('chat'))
        # distinguish disabled / locked / bad-credential (match username OR email,
        # so the right message shows even when the person signs in with their email)
        row = users_db.get_by_identifier(username)
        if row and not row.get('is_active', 1):
            error = 'This account is disabled. Contact an administrator.'
        elif row and users_db.is_locked(row):
            error = 'Too many failed attempts. Try again in ~15 minutes.'
        else:
            error = 'Invalid username or password.'
    return render_template('login.html', error=error)

@chatbot_app.route('/home')
def home():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    return render_template('home.html', username=session.get('username'),
                           is_admin=session.get('is_admin', False), active='home',
                           perms=_perms())


@chatbot_app.route('/chat')
@require_module('chat')
def chat():
    return render_template('chat.html', username=session.get('username'),
                           is_admin=session.get('is_admin', False), active='chat',
                           perms=_perms())

@chatbot_app.route('/ask', methods=['POST'])
def ask():
    """Non-streaming fallback — full JSON answer in one response."""
    if not session.get('logged_in'):
        return jsonify({'error': 'Not logged in'}), 401
    user_question = (request.json or {}).get('question', '').strip()
    if not user_question:
        return jsonify({'error': 'No question provided'}), 400

    t0 = time.time()
    username = session.get('username')
    history = _sanitize_history((request.json or {}).get('history'))
    result, in_tok, out_tok, retries = _resolve_and_run(user_question, history)
    kind = result["kind"]

    if kind == "unable":
        _log_ask(username, user_question, t0, ok=False, sql='', retries=retries, in_tok=in_tok, out_tok=out_tok)
        return jsonify({'answer': "I'm unable to answer that from the inventory data I have access to.", 'sql': ''})
    if kind == "error":
        _log_ask(username, user_question, t0, ok=False, sql=result["sql"], error=result["error"],
                 retries=retries, in_tok=in_tok, out_tok=out_tok)
        return jsonify({'answer': FRIENDLY_ERROR, 'sql': result["sql"]})
    if kind == "empty":
        _log_ask(username, user_question, t0, ok=True, sql=result["sql"], row_count=0,
                 retries=retries, in_tok=in_tok, out_tok=out_tok)
        return jsonify({'answer': 'No results found for your question.', 'sql': result["sql"],
                        'columns': result["columns"], 'rows': []})

    sql, data = result["sql"], result["data"]
    answer = format_answer(sql, data, user_question, truncated=result["truncated"], total_rows=result["total_rows"])
    _log_ask(username, user_question, t0, ok=True, sql=sql, row_count=result["total_rows"],
             retries=retries, in_tok=in_tok, out_tok=out_tok)
    return jsonify({'answer': answer, 'sql': sql, 'rows': data['rows'][:CHAT_ROW_CAP],
                    'columns': data['columns'], 'truncated': result["truncated"], 'total_rows': result["total_rows"]})


@chatbot_app.route('/ask/stream', methods=['POST'])
def ask_stream():
    """Streaming answer over newline-delimited JSON events: stage → meta → delta → done."""
    if not session.get('logged_in'):
        return jsonify({'error': 'Not logged in'}), 401
    user_question = (request.json or {}).get('question', '').strip()
    if not user_question:
        return jsonify({'error': 'No question provided'}), 400
    history = _sanitize_history((request.json or {}).get('history'))
    username = session.get('username')

    def ev(**kw):
        return json.dumps(kw, default=_ndjson_default) + "\n"

    def gen():
        t0 = time.time()
        in_tok = out_tok = retries = 0
        result = None
        sql = None
        try:
            for item in _resolve_stream(user_question, history):
                if "stage" in item:
                    yield ev(type="stage", stage=item["stage"])
                else:
                    result = item["result"]
                    in_tok, out_tok, retries = item["in_tok"], item["out_tok"], item["retries"]
            kind = result["kind"]
            if kind == "unable":
                yield ev(type="final", answer="I'm unable to answer that from the inventory data I have access to.", sql="")
                yield ev(type="done")
                _log_ask(username, user_question, t0, ok=False, sql='', retries=retries, in_tok=in_tok, out_tok=out_tok)
                return
            if kind == "error":
                sql = result["sql"]
                yield ev(type="final", answer=FRIENDLY_ERROR, sql=sql, error=True)
                yield ev(type="done")
                _log_ask(username, user_question, t0, ok=False, sql=sql, error=result["error"],
                         retries=retries, in_tok=in_tok, out_tok=out_tok)
                return
            if kind == "empty":
                sql = result["sql"]
                yield ev(type="meta", sql=sql, columns=result["columns"], rows=[], truncated=False, total_rows=0)
                yield ev(type="final", answer="No results found for your question.", sql=sql)
                yield ev(type="done")
                _log_ask(username, user_question, t0, ok=True, sql=sql, row_count=0,
                         retries=retries, in_tok=in_tok, out_tok=out_tok)
                return

            sql, data = result["sql"], result["data"]
            yield ev(type="meta", sql=sql, columns=data['columns'], rows=data['rows'][:CHAT_ROW_CAP],
                     truncated=result["truncated"], total_rows=result["total_rows"])
            yield ev(type="stage", stage="summarizing")
            usage_out = {}
            for chunk in format_answer_stream(sql, data, user_question, result["truncated"],
                                              result["total_rows"], usage_out):
                yield ev(type="delta", text=chunk)
            in_tok += usage_out.get('input_tokens', 0)
            out_tok += usage_out.get('output_tokens', 0)
            yield ev(type="done")
            _log_ask(username, user_question, t0, ok=True, sql=sql, row_count=result["total_rows"],
                     retries=retries, in_tok=in_tok, out_tok=out_tok)
        except Exception as e:
            try:
                yield ev(type="error", message="Something went wrong. Please try again.")
            except Exception:
                pass
            _log_ask(username, user_question, t0, ok=False, sql=sql, error=str(e),
                     retries=retries, in_tok=in_tok, out_tok=out_tok)

    return Response(gen(), mimetype="application/x-ndjson",
                    headers={"X-Accel-Buffering": "no", "Cache-Control": "no-cache"})


def _ndjson_default(o):
    """JSON fallback for stream events: DB Decimal -> float, dates -> ISO string.
    (Plain json.dumps can't serialize Decimal/datetime, which broke any query with a
    cost/price/date column.)"""
    import datetime as _dt
    from decimal import Decimal
    if isinstance(o, Decimal):
        return float(o)
    if isinstance(o, (_dt.datetime, _dt.date, _dt.time)):
        return o.isoformat()
    return str(o)


def _xl(v):
    """Coerce a DB value to something openpyxl can write."""
    import datetime as _dt
    from decimal import Decimal
    if v is None or isinstance(v, (str, int, float, bool, Decimal, _dt.datetime, _dt.date, _dt.time)):
        return v
    return str(v)


# Neutralize spreadsheet formula injection: a *string* cell that begins with a
# formula trigger gets a leading apostrophe so Excel/Sheets treat it as text.
_FORMULA_TRIGGERS = ('=', '+', '-', '@', '\t', '\r')


def _safe_cell(v):
    if isinstance(v, str) and v.startswith(_FORMULA_TRIGGERS):
        return "'" + v
    return v


@chatbot_app.route('/ask/export', methods=['POST'])
def ask_export():
    """Download the current query's full results as Excel or CSV (SQL re-validated,
    SELECT-only over the single allowed table)."""
    if not session.get('logged_in'):
        return jsonify({'error': 'Not logged in'}), 401
    body = request.json or {}
    sql = (body.get('sql') or '').strip()
    fmt = (body.get('format') or 'xlsx').lower()
    if not sql:
        return jsonify({'error': 'No query to export'}), 400
    try:
        safe_sql = chat_sql.validate_sql(sql)
    except chat_sql.SqlValidationError as e:
        return jsonify({'error': f'Query not allowed: {e}'}), 400
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(safe_sql)
        cols = [c[0] for c in cur.description]
        rows = [list(r) for r in cur.fetchmany(EXPORT_ROW_CAP)]
        conn.close()
    except Exception:
        return jsonify({'error': 'Could not run the query.'}), 500

    if fmt == 'csv':
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow([_safe_cell(c) for c in cols])
        w.writerows([[_safe_cell(v) for v in r] for r in rows])
        return Response(buf.getvalue().encode('utf-8-sig'), mimetype='text/csv',
                        headers={'Content-Disposition': 'attachment; filename="inventory-export.csv"'})

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.append([_safe_cell(c) for c in cols])
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
    for r in rows:
        ws.append([_safe_cell(_xl(v)) for v in r])
    out = io.BytesIO()
    wb.save(out)
    return Response(out.getvalue(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': 'attachment; filename="inventory-export.xlsx"'})

@chatbot_app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# --- Set Password (invite flow) ---
@chatbot_app.route('/set-password/<token>', methods=['GET', 'POST'])
def set_password(token):
    user = users_db.get_user_by_token(token)
    if not user:
        return render_template('set_password.html', error='This invite link is invalid or has expired.', token=None, username=None)

    if request.method == 'POST':
        password = request.form.get('password', '').strip()
        confirm = request.form.get('confirm', '').strip()
        if len(password) < 6:
            return render_template('set_password.html', error='Password must be at least 6 characters.', token=token, username=user['username'])
        if password != confirm:
            return render_template('set_password.html', error='Passwords do not match.', token=token, username=user['username'])
        users_db.set_password_by_token(token, password)
        return render_template('set_password.html', success=True, token=None, username=user['username'])

    return render_template('set_password.html', token=token, username=user['username'], error=None)


# --- Admin: User Management ---
@chatbot_app.route('/admin/users')
def admin_users():
    if not session.get('logged_in') or not session.get('is_admin'):
        return redirect(url_for('login'))
    users = users_db.get_all_users()
    return render_template('admin_users.html', users=users,
                           username=session.get('username'),
                           is_admin=session.get('is_admin', False), active='admin')


@chatbot_app.route('/admin/chat-log')
def admin_chat_log():
    if not session.get('logged_in') or not session.get('is_admin'):
        return redirect(url_for('login'))
    return render_template('chat_log.html', logs=chat_log.recent(200),
                           username=session.get('username'),
                           is_admin=session.get('is_admin', False), active='admin')


@chatbot_app.route('/admin/audit')
def admin_audit_view():
    if not session.get('logged_in') or not session.get('is_admin'):
        return redirect(url_for('login'))
    return render_template('admin_audit.html', logs=admin_audit.recent(200),
                           username=session.get('username'),
                           is_admin=True, active='audit')


@chatbot_app.route('/admin/users/create', methods=['POST'])
def admin_create_user():
    if not session.get('logged_in') or not session.get('is_admin'):
        return jsonify({'ok': False, 'error': 'Unauthorized'}), 403
    data = request.get_json()
    new_username = (data.get('username') or '').strip()
    new_email = (data.get('email') or '').strip()
    is_admin = bool(data.get('is_admin', False))
    if not new_username or not new_email:
        return jsonify({'ok': False, 'error': 'Username and email are required'})
    try:
        token = users_db.create_user(new_username, new_email, is_admin,
                                     created_by=session.get('username'))
        send_invite_email(new_email, new_username, token)
        admin_audit.log_action(session.get('username'), 'create_user',
                               target=new_username, detail=new_email)
        return jsonify({'ok': True})
    except Exception as e:
        if 'UNIQUE' in str(e).upper():
            return jsonify({'ok': False, 'error': f'Username "{new_username}" already exists'})
        return jsonify({'ok': False, 'error': str(e)})


@chatbot_app.route('/admin/users/resend-invite', methods=['POST'])
def admin_resend_invite():
    if not session.get('logged_in') or not session.get('is_admin'):
        return jsonify({'ok': False, 'error': 'Unauthorized'}), 403
    data = request.get_json()
    user_id = data.get('id')
    user = users_db.get_user_by_id(user_id)
    if not user:
        return jsonify({'ok': False, 'error': 'User not found'})
    if not user.get('email'):
        return jsonify({'ok': False, 'error': 'User has no email address'})
    try:
        token = users_db.generate_invite_token(user_id)
        send_invite_email(user['email'], user['username'], token)
        admin_audit.log_action(session.get('username'), 'resend_invite',
                               target=user['username'])
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


@chatbot_app.route('/admin/users/reset-password', methods=['POST'])
def admin_reset_password():
    if not session.get('logged_in') or not session.get('is_admin'):
        return jsonify({'ok': False, 'error': 'Unauthorized'}), 403
    data = request.get_json()
    user_id = data.get('id')
    user = users_db.get_user_by_id(user_id)
    if not user:
        return jsonify({'ok': False, 'error': 'User not found'})
    if not user.get('email'):
        return jsonify({'ok': False, 'error': 'User has no email address'})
    try:
        token = users_db.generate_invite_token(user_id)
        send_invite_email(user['email'], user['username'], token)
        admin_audit.log_action(session.get('username'), 'reset_password',
                               target=user['username'])
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


@chatbot_app.route('/admin/users/toggle-admin', methods=['POST'])
def admin_toggle_admin():
    if not session.get('logged_in') or not session.get('is_admin'):
        return jsonify({'ok': False, 'error': 'Unauthorized'}), 403
    data = request.get_json()
    user_id = data.get('id')
    user = users_db.get_user_by_id(user_id)
    if not user:
        return jsonify({'ok': False, 'error': 'User not found'})
    if user['username'] == session.get('username'):
        return jsonify({'ok': False, 'error': 'Cannot change your own admin status'})
    new_admin_state = not user['is_admin']
    users_db.update_admin_status(user_id, new_admin_state)
    admin_audit.log_action(session.get('username'), 'toggle_admin',
                           target=user['username'],
                           detail='admin' if new_admin_state else 'not_admin')
    return jsonify({'ok': True, 'is_admin': new_admin_state})


@chatbot_app.route('/admin/users/set-role', methods=['POST'])
def admin_set_role():
    if not session.get('logged_in') or not session.get('is_admin'):
        return jsonify({'ok': False, 'error': 'Unauthorized'}), 403
    d = request.get_json() or {}
    user = users_db.get_user_by_id(d.get('id'))
    if not user:
        return jsonify({'ok': False, 'error': 'User not found'})
    if user['username'] == session.get('username'):
        return jsonify({'ok': False, 'error': 'Cannot change your own role'})
    role = d.get('role')
    if role not in roles.ROLES:
        return jsonify({'ok': False, 'error': 'Invalid role'})
    users_db.set_role(d['id'], role)
    admin_audit.log_action(session.get('username'), 'set_role', target=user['username'], detail=role)
    return jsonify({'ok': True})


@chatbot_app.route('/admin/users/delete', methods=['POST'])
def admin_delete_user():
    if not session.get('logged_in') or not session.get('is_admin'):
        return jsonify({'ok': False, 'error': 'Unauthorized'}), 403
    data = request.get_json()
    user_id = data.get('id')
    user = users_db.get_user_by_id(user_id)
    if not user:
        return jsonify({'ok': False, 'error': 'User not found'})
    if user['username'] == session.get('username'):
        return jsonify({'ok': False, 'error': 'Cannot delete your own account'})
    users_db.delete_user(user_id)
    admin_audit.log_action(session.get('username'), 'delete_user', target=user['username'])
    return jsonify({'ok': True})


@chatbot_app.route('/admin/users/edit', methods=['POST'])
def admin_edit_user():
    if not session.get('logged_in') or not session.get('is_admin'):
        return jsonify({'ok': False, 'error': 'Unauthorized'}), 403
    d = request.get_json() or {}
    user = users_db.get_user_by_id(d.get('id'))
    if not user:
        return jsonify({'ok': False, 'error': 'User not found'})
    new_username = (d.get('username') or '').strip()
    new_email = (d.get('email') or '').strip()
    if not new_username:
        return jsonify({'ok': False, 'error': 'Username required'})
    is_self = user['username'] == session.get('username')
    actor = session.get('username')
    try:
        if new_username != user['username']:
            users_db.update_username(d['id'], new_username)
        users_db.set_email(d['id'], new_email)
    except Exception as e:
        if 'UNIQUE' in str(e).upper():
            return jsonify({'ok': False, 'error': f'Username "{new_username}" already exists'})
        return jsonify({'ok': False, 'error': str(e)})
    # Apply role change (non-self only)
    new_role = d.get('role')
    if new_role is not None and not is_self:
        if new_role not in roles.ROLES:
            return jsonify({'ok': False, 'error': 'Invalid role'})
        if new_role != user.get('role'):
            users_db.set_role(d['id'], new_role)
            admin_audit.log_action(actor, 'set_role', target=new_username, detail=new_role)
    # Apply active change (non-self only)
    if 'active' in d and not is_self:
        raw_active = d['active']
        if isinstance(raw_active, bool):
            new_active = raw_active
        else:
            new_active = str(raw_active).lower() not in ('false', '0', '')
        users_db.set_active(d['id'], new_active)
        admin_audit.log_action(actor, 'set_active', target=new_username,
                               detail='active' if new_active else 'disabled')
    admin_audit.log_action(actor, 'edit_user', target=new_username,
                           detail=f"email={new_email}")
    return jsonify({'ok': True})


@chatbot_app.route('/admin/users/set-active', methods=['POST'])
def admin_set_active():
    if not session.get('logged_in') or not session.get('is_admin'):
        return jsonify({'ok': False, 'error': 'Unauthorized'}), 403
    d = request.get_json() or {}
    user = users_db.get_user_by_id(d.get('id'))
    if not user:
        return jsonify({'ok': False, 'error': 'User not found'})
    if user['username'] == session.get('username'):
        return jsonify({'ok': False, 'error': 'Cannot disable your own account'})
    active = bool(d.get('active'))
    users_db.set_active(d['id'], active)
    admin_audit.log_action(session.get('username'), 'set_active', target=user['username'],
                           detail='active' if active else 'disabled')
    return jsonify({'ok': True})


# --- Profile: self-service password and email ---
@chatbot_app.route('/profile')
def profile():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    me = users_db._row_by_username(session.get('username'))
    return render_template('profile.html', me=me,
                           username=session.get('username'),
                           is_admin=session.get('is_admin', False),
                           active='profile',
                           perms=_perms())


@chatbot_app.route('/profile/password', methods=['POST'])
def profile_password():
    if not session.get('logged_in'):
        return jsonify({'ok': False, 'error': 'Not logged in'}), 401
    d = request.get_json() or {}
    if not users_db.verify_password(session['username'], d.get('current', '')):
        return jsonify({'ok': False, 'error': 'Current password is incorrect'})
    if len(d.get('new', '')) < 6:
        return jsonify({'ok': False, 'error': 'New password must be at least 6 characters'})
    uid = users_db._row_by_username(session['username'])['id']
    users_db.update_password(uid, d['new'])
    return jsonify({'ok': True})


@chatbot_app.route('/profile/email', methods=['POST'])
def profile_email():
    if not session.get('logged_in'):
        return jsonify({'ok': False, 'error': 'Not logged in'}), 401
    d = request.get_json() or {}
    uid = users_db._row_by_username(session['username'])['id']
    users_db.set_email(uid, (d.get('email') or '').strip())
    return jsonify({'ok': True})


app = chatbot_app

if __name__ == '__main__':
    chatbot_app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)), debug=False)
