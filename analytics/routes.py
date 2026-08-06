from flask import Blueprint, request, jsonify, session, redirect, url_for, Response
from analytics import db, pricing, templates
from io import BytesIO
from datetime import datetime
import admin_audit
import roles

analytics_bp = Blueprint('analytics', __name__, url_prefix='/analytics')


@analytics_bp.before_request
def _gate_analytics():
    role = roles.effective_role(session.get('role'), session.get('is_admin'))
    if session.get('logged_in') and not roles.role_allows(role, 'analytics'):
        return redirect(url_for('home'))


def _require_login():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    return None


@analytics_bp.route('/')
def index():
    redir = _require_login()
    if redir:
        return redir
    return templates.render_analytics_index()


def _telus_options():
    """Fetch the combobox option lists (tags + client names). Every form render —
    including error re-renders — must pass these, or the comboboxes come back empty
    and unusable ("No matches" for everything)."""
    try:
        project_tags = db.get_telus_project_tags()
    except Exception:
        project_tags = []
    try:
        client_names = db.get_telus_client_names()
    except Exception:
        client_names = []
    return project_tags, client_names


@analytics_bp.route('/telus-weekly')
def telus_weekly_form():
    redir = _require_login()
    if redir:
        return redir
    project_tags, client_names = _telus_options()
    return templates.render_telus_weekly_form(
        project_tags=project_tags,
        client_names=client_names,
    )


@analytics_bp.route('/telus-weekly/client-for-tag')
def telus_weekly_client_for_tag():
    """AJAX: distinct client name(s) for a ProjectTag. Returns `client` (a single
    name) only when the tag maps unambiguously to one client, so the form can
    auto-fill it; ambiguous/none -> `client` is null and the field is left alone."""
    if not session.get('logged_in'):
        return jsonify({'ok': False, 'error': 'Not logged in'}), 401
    tag = request.args.get('project_tag', '').strip()
    if not tag:
        return jsonify({'ok': True, 'clients': [], 'client': None})
    try:
        clients = db.get_client_for_project_tag(tag)
    except Exception:
        clients = []
    return jsonify({'ok': True, 'clients': clients,
                    'client': clients[0] if len(clients) == 1 else None})


@analytics_bp.route('/telus-weekly/report', methods=['POST'])
def telus_weekly_report():
    redir = _require_login()
    if redir:
        return redir

    project_tag = request.form.get('project_tag', '').strip()
    client_name = request.form.get('client_name', '').strip() or None
    admin_audit.stash(action='telus_report', target=project_tag, client=client_name)

    if not project_tag:
        project_tags, client_names = _telus_options()
        return templates.render_telus_weekly_form(
            error='ProjectTag is required.',
            project_tags=project_tags, client_names=client_names)

    try:
        devices = db.call_repair_assessment(project_tag, client_name)
    except Exception as e:
        admin_audit.stash(result='db_error', error=str(e)[:300])
        project_tags, client_names = _telus_options()
        return templates.render_telus_weekly_form(
            error=f'Database error: {e}',
            project_tag=project_tag,
            client_name=client_name,
            project_tags=project_tags, client_names=client_names)

    if not devices:
        admin_audit.stash(result='no_devices')
        project_tags, client_names = _telus_options()
        return templates.render_telus_weekly_form(
            error=f'No devices found for ProjectTag "{project_tag}".',
            project_tag=project_tag,
            client_name=client_name,
            project_tags=project_tags, client_names=client_names)

    pricing_map = db.get_pricing_map()
    enriched, summary = pricing.compute_report(devices, pricing_map)
    admin_audit.stash(device_count=summary['total_devices'], total_lot_value=summary['total_lot_value'])

    return templates.render_telus_weekly_report(
        project_tag, client_name, enriched, summary,
    )


@analytics_bp.route('/telus-weekly/export', methods=['POST'])
def telus_weekly_export():
    redir = _require_login()
    if redir:
        return redir

    project_tag = request.form.get('project_tag', '').strip()
    client_name = request.form.get('client_name', '').strip() or None
    admin_audit.stash(action='telus_export', target=project_tag, client=client_name)

    if not project_tag:
        return redirect(url_for('analytics.telus_weekly_form'))

    devices = db.call_repair_assessment(project_tag, client_name)
    pricing_map = db.get_pricing_map()
    enriched, summary = pricing.compute_report(devices, pricing_map)
    admin_audit.stash(device_count=len(enriched))

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()

    # --- Repair & Resell sheet ---
    ws = wb.active
    ws.title = 'Repair & Resell'

    headers = [
        'ESN', 'Origin', 'Make', 'Model', 'Memory', 'Condition',
        'Fault 1', 'Fault 2', 'Fault 3', 'QC Notes',
        'Unassessed Price', 'Received Grade', 'Assessed Price',
        'Repair Labour', 'Repair Parts', 'Parts Used', 'Total Repair Cost',
        'Grade After Repair', 'Price After Repair', 'Upside',
        'Grade Improvement', 'Improvement Labour', 'Improvement Parts',
        'Total Improvement', 'Grade After Improvement',
        'Total Repair + Improvement', 'Price After Improvement',
        'Improvement Upside', 'Recommendation', 'Lot Value',
    ]

    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    def _val(v):
        if v is None:
            return 'N/A'
        return v

    for row_idx, row in enumerate(enriched, 2):
        vals = [
            row.get('ESN'),
            row.get('Vendor'),
            row.get('ManufacturerVerb'),
            row.get('ModelVerb'),
            row.get('Memory'),
            row.get('Conditions'),
            row.get('Defects_1'),
            row.get('Defects_2'),
            row.get('Defects_3'),
            row.get('QC_Notes'),
            row.get('unassessed_price'),
            row.get('Received_Grade'),
            row.get('assessed_price'),
            row.get('T_Level_Cost'),
            row.get('T_Part_Cost'),
            row.get('Parts_Used'),
            row.get('total_repair_cost'),
            row.get('Post-Repair_Grade'),
            row.get('price_after_repair'),
            row.get('upside'),
            row.get('Grade_Improvement'),
            row.get('T_Level_Improved_Cos'),
            row.get('T_Part_Improved_Cost'),
            row.get('total_improvement_cost'),
            row.get('Post_Improved_Grade'),
            row.get('total_repair_plus_improvement'),
            row.get('price_after_improvement'),
            row.get('improvement_upside'),
            row.get('recommendation'),
            row.get('lot_value'),
        ]
        for col_idx, v in enumerate(vals, 1):
            ws.cell(row=row_idx, column=col_idx, value=_val(v))

    # --- Models sheet ---
    ws_models = wb.create_sheet('Models')
    model_counts = {}
    for row in enriched:
        m = row.get('ModelVerb') or 'Unknown'
        model_counts[m] = model_counts.get(m, 0) + 1

    ws_models.cell(row=1, column=1, value='Model').font = Font(bold=True)
    ws_models.cell(row=1, column=2, value='Count').font = Font(bold=True)
    for i, (model, cnt) in enumerate(sorted(model_counts.items()), 2):
        ws_models.cell(row=i, column=1, value=model)
        ws_models.cell(row=i, column=2, value=cnt)

    # --- Summary sheet ---
    ws_summary = wb.create_sheet('Summary')
    ws_summary.cell(row=1, column=1, value='Total Devices').font = Font(bold=True)
    ws_summary.cell(row=1, column=2, value=summary['total_devices'])
    ws_summary.cell(row=2, column=1, value='Total Lot Value').font = Font(bold=True)
    ws_summary.cell(row=2, column=2, value=summary['total_lot_value'])
    r = 4
    ws_summary.cell(row=r, column=1, value='Recommendations').font = Font(bold=True)
    for rec, cnt in summary['recommendation_breakdown'].items():
        r += 1
        ws_summary.cell(row=r, column=1, value=rec)
        ws_summary.cell(row=r, column=2, value=cnt)
    r += 2
    ws_summary.cell(row=r, column=1, value='Conditions').font = Font(bold=True)
    for cond, cnt in summary['conditions_breakdown'].items():
        r += 1
        ws_summary.cell(row=r, column=1, value=cond)
        ws_summary.cell(row=r, column=2, value=cnt)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    today = datetime.now().strftime('%Y%m%d')
    filename = f'TW_{project_tag}_{today}.xlsx'
    admin_audit.stash(filename=filename)

    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'},
    )


# ---------------------------------------------------------------------------
# Price Review
# ---------------------------------------------------------------------------

def _float_prices(models):
    for m in models:
        for field in ('GradeA_Price', 'GradeB_Price', 'GradeC_Price',
                      'Defective_Price', 'FRP_Price'):
            if m.get(field) is not None:
                m[field] = float(m[field])


@analytics_bp.route('/price-review')
def price_review():
    redir = _require_login()
    if redir:
        return redir

    project_tag = request.args.get('project_tag', '').strip()

    if project_tag:
        try:
            model_counts = db.get_distinct_models_for_project(project_tag)
        except Exception as e:
            models = db.get_all_pricing_models()
            _float_prices(models)
            return templates.render_price_review(
                models, error=f'Error loading project: {e}')

        if not model_counts:
            models = db.get_all_pricing_models()
            _float_prices(models)
            return templates.render_price_review(
                models,
                error=f'No devices found for ProjectTag "{project_tag}".')

        existing = db.get_pricing_models_by_names(list(model_counts.keys()))
        _float_prices(existing)

        existing_lower = {m['Model'].strip().lower() for m in existing}
        for m in existing:
            key = m['Model'].strip().lower()
            for k, v in model_counts.items():
                if k.strip().lower() == key:
                    m['device_count'] = v
                    break

        new_models = []
        for name, count in sorted(model_counts.items()):
            if name.strip().lower() not in existing_lower:
                new_models.append({'Model': name, 'count': count})

        total_devices = sum(model_counts.values())
        return templates.render_price_review(
            existing, project_tag=project_tag, new_models=new_models,
            total_project_devices=total_devices,
            total_project_models=len(model_counts))

    models = db.get_all_pricing_models()
    _float_prices(models)
    return templates.render_price_review(models)



# maps a price-review update's field name -> the TelusWeeklyPricingMaster
# column returned by db.get_all_pricing_models(), for the old->new audit diff.
_PRICE_UPDATE_FIELDS = {
    'grade_a':     'GradeA_Price',
    'grade_b':     'GradeB_Price',
    'grade_c':     'GradeC_Price',
    'defective':   'Defective_Price',
    'frp':         'FRP_Price',
    'device_type': 'DeviceType',
}


def _price_update_changes(updates):
    """Old->new diff for the price_update audit stash, keyed by model name,
    capped at 20 models + 'more': N. Only fields that actually changed are
    included. Raises on DB error — the caller falls back to new-values-only
    so a failed audit lookup never blocks the save."""
    ids = {u.get('id') for u in updates if u.get('id') is not None}
    old_by_id = {row['ID']: row for row in db.get_all_pricing_models() if row.get('ID') in ids}
    changes = {}
    for u in updates:
        old = old_by_id.get(u.get('id'))
        if not old:
            continue
        model_name = old.get('Model') or str(u.get('id'))
        field_changes = {}
        for key, col in _PRICE_UPDATE_FIELDS.items():
            new_val = u.get(key)
            old_val = old.get(col)
            if key != 'device_type':
                old_val = float(old_val) if old_val is not None else None
                new_val = float(new_val) if new_val is not None else None
            if old_val != new_val:
                field_changes[key] = {'old': old_val, 'new': new_val}
        if field_changes:
            changes[model_name] = field_changes
    if len(changes) > 20:
        capped = dict(list(changes.items())[:20])
        capped['more'] = len(changes) - 20
        return capped
    return changes


@analytics_bp.route('/price-review/save', methods=['POST'])
def price_review_save():
    if not session.get('logged_in'):
        return jsonify({'ok': False, 'error': 'Not logged in'}), 401

    data = request.get_json()
    updates = data.get('updates', [])
    if not updates:
        return jsonify({'ok': False, 'error': 'No updates provided'})

    try:
        changes = _price_update_changes(updates)
        admin_audit.stash(action='price_update', count=len(updates), changes=changes)
    except Exception:
        admin_audit.stash(action='price_update', count=len(updates), values=updates[:20])

    try:
        db.bulk_update_pricing(updates, updated_by=session.get('username'))
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


@analytics_bp.route('/price-review/bulk-add', methods=['POST'])
def price_review_bulk_add():
    if not session.get('logged_in'):
        return jsonify({'ok': False, 'error': 'Not logged in'}), 401

    data = request.get_json()
    models = data.get('models', [])
    if not models:
        return jsonify({'ok': False, 'error': 'No models provided'})

    admin_audit.stash(action='price_bulk_add', count=len(models),
                      models=[(m.get('model') or '')[:80] for m in models][:20])

    try:
        new_ids = db.bulk_insert_pricing_models(models)
        return jsonify({'ok': True, 'count': len(new_ids)})
    except Exception as e:
        if 'UNIQUE' in str(e).upper() or 'duplicate' in str(e).lower():
            return jsonify({'ok': False,
                            'error': 'One or more models already exist'})
        return jsonify({'ok': False, 'error': str(e)})


@analytics_bp.route('/price-review/add', methods=['POST'])
def price_review_add():
    if not session.get('logged_in'):
        return jsonify({'ok': False, 'error': 'Not logged in'}), 401

    data = request.get_json()
    model = (data.get('model') or '').strip()
    if not model:
        return jsonify({'ok': False, 'error': 'Model name required'})

    grade_a = float(data.get('grade_a', 0))
    grade_b = float(data.get('grade_b', 0))
    grade_c = float(data.get('grade_c', 0))
    defective = float(data.get('defective', 0))
    frp = float(data.get('frp', 0))
    device_type = data.get('device_type', 'Phone')
    admin_audit.stash(action='price_add', target=model,
                      prices={'grade_a': grade_a, 'grade_b': grade_b, 'grade_c': grade_c,
                              'defective': defective, 'frp': frp, 'device_type': device_type})

    try:
        new_id = db.insert_pricing_model(
            model, grade_a, grade_b, grade_c, defective, frp, device_type,
        )
        return jsonify({'ok': True, 'id': new_id})
    except Exception as e:
        if 'UNIQUE' in str(e).upper() or 'duplicate' in str(e).lower():
            return jsonify({'ok': False, 'error': f'Model "{model}" already exists'})
        return jsonify({'ok': False, 'error': str(e)})
